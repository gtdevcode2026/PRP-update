import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

REQUIRED_SHEET = "OneTrust Assessment"
REQUIRED_COLUMNS = ["ID", "Stage", "Organization", "Ageing"]
STAGE_ORDER = ["Completed", "In progress", "Not started", "Under review"]
ORG_ORDER = [
    "Africa",
    "APAC",
    "BEES",
    "BEES | FINTECH",
    "Europe",
    "GHQ",
    "Middle America Zone",
    "North America Zone",
    "South America Zone",
]
ZONE_MAP = {
    "Africa": "AFR",
    "APAC": "APAC",
    "BEES": "GRO",
    "BEES | FINTECH": "GRO",
    "Europe": "EUR",
    "GHQ": "GHQ",
    "Middle America Zone": "MAZ",
    "North America Zone": "NAZ",
    "South America Zone": "SAZ",
}
ZONE_ORDER = ["AFR", "APAC", "GRO", "EUR", "GHQ", "MAZ", "NAZ", "SAZ"]
DEFAULT_OVERDUE_THRESHOLD = 90

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="EAF3FB")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def autosize_columns(ws, min_width=12, max_width=32):
    for col_cells in ws.columns:
        lengths = []
        for c in col_cells:
            try:
                lengths.append(len(str(c.value)) if c.value is not None else 0)
            except Exception:
                lengths.append(0)
        width = max(lengths) + 2 if lengths else min_width
        width = max(min_width, min(max_width, width))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def style_table(ws, start_row, start_col, end_row, end_col, grand_total_row=None):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            cell.border = THIN_BORDER
            if r == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
            elif grand_total_row and r == grand_total_row:
                cell.fill = SUBHEADER_FILL
                cell.font = HEADER_FONT
    for r in range(start_row + 1, end_row + 1):
        ws.cell(r, start_col).font = Font(bold=(grand_total_row == r))


def add_stacked_chart(ws, top_left_cell, title, data_start_row, data_end_row, category_col=1, min_data_col=2, max_data_col=3):
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = title
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Zone"
    chart.height = 8
    chart.width = 16
    chart.legend.position = "b"

    data_ref = Reference(ws, min_col=min_data_col, max_col=max_data_col, min_row=data_start_row, max_row=data_end_row)
    cats_ref = Reference(ws, min_col=category_col, min_row=data_start_row + 1, max_row=data_end_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = False

    try:
        chart.series[0].graphicalProperties.solidFill = "1F77B4"  # Closed
        chart.series[0].graphicalProperties.line.solidFill = "1F77B4"
        chart.series[1].graphicalProperties.solidFill = "FF7F0E"  # Open
        chart.series[1].graphicalProperties.line.solidFill = "FF7F0E"
    except Exception:
        pass

    ws.add_chart(chart, top_left_cell)


def build_reports(input_file, output_file=None, overdue_threshold=DEFAULT_OVERDUE_THRESHOLD):
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    if output_file is None:
        output_file = input_path.with_name(f"{input_path.stem}_report_only.xlsx")
    else:
        output_file = Path(output_file)

    # Read only the source data sheet from the input workbook
    df = pd.read_excel(input_path, sheet_name=REQUIRED_SHEET, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns in '{REQUIRED_SHEET}': {missing}")

    data = df[REQUIRED_COLUMNS].copy()
    data["Stage"] = data["Stage"].astype(str).str.strip()
    data["Organization"] = data["Organization"].astype(str).str.strip()
    data["Ageing"] = pd.to_numeric(data["Ageing"], errors="coerce")
    data = data[data["Stage"].isin(STAGE_ORDER)]
    data = data[data["Organization"].isin(ORG_ORDER)]
    data["Zone"] = data["Organization"].map(ZONE_MAP)
    data["Is Overdue"] = data["Ageing"] >= overdue_threshold

    # Pivot-style organization summary
    pivot_df = pd.pivot_table(data, index="Organization", columns="Stage", values="ID", aggfunc="count", fill_value=0)
    pivot_df = pivot_df.reindex(index=ORG_ORDER, columns=STAGE_ORDER, fill_value=0)
    pivot_df["Grand Total"] = pivot_df.sum(axis=1)
    pivot_df.loc["Grand Total"] = pivot_df.sum(axis=0)

    # Zone summary
    zone_stage = pivot_df.drop(index="Grand Total").copy().reset_index()
    zone_stage["Zone"] = zone_stage["Organization"].map(ZONE_MAP)
    zone_summary = zone_stage.groupby("Zone", as_index=True).agg({
        "Completed": "sum",
        "In progress": "sum",
        "Not started": "sum",
        "Under review": "sum",
        "Grand Total": "sum",
    })
    zone_summary["Closed"] = zone_summary["Completed"] + zone_summary["Under review"]
    zone_summary["Open"] = zone_summary["In progress"] + zone_summary["Not started"]
    zone_summary = zone_summary.reindex(ZONE_ORDER)
    zone_summary = zone_summary[["Closed", "Open", "Grand Total"]]
    total_open = int(zone_summary["Open"].sum())
    total_all = int(zone_summary["Grand Total"].sum())

    # Overdue summary
    overdue = data[data["Is Overdue"]].copy()
    overdue_zone = overdue.groupby(["Zone", "Stage"]).size().unstack(fill_value=0)
    overdue_zone = overdue_zone.reindex(index=ZONE_ORDER, columns=STAGE_ORDER, fill_value=0)
    overdue_zone["Closed"] = overdue_zone["Completed"] + overdue_zone["Under review"]
    overdue_zone["Open"] = overdue_zone["In progress"] + overdue_zone["Not started"]
    overdue_zone["Grand Total"] = overdue_zone["Closed"] + overdue_zone["Open"]
    total_overdue = int(overdue_zone["Grand Total"].sum())
    total_overdue_open = int(overdue_zone["Open"].sum())

    # IMPORTANT: create a brand-new workbook so output contains ONLY the generated report sheets
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    ws1 = wb.create_sheet("Auto Pivot Summary")
    ws2 = wb.create_sheet("Auto Open Closed")
    ws3 = wb.create_sheet(f"Auto Overdue {overdue_threshold}D")

    # Sheet 1
    ws1["A1"] = "Pivot-style summary from 'OneTrust Assessment'"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A3"] = "Fields used"
    ws1["A3"].font = Font(bold=True)
    ws1["A4"] = "Rows"; ws1["B4"] = "Organization"
    ws1["A5"] = "Columns"; ws1["B5"] = "Stage"
    ws1["A6"] = "Values"; ws1["B6"] = "Count of ID"

    start_row = 8
    headers = ["Organization"] + list(pivot_df.columns)
    for col_num, header in enumerate(headers, start=1):
        ws1.cell(start_row, col_num, header)
    for row_num, (idx, row) in enumerate(pivot_df.iterrows(), start=start_row + 1):
        ws1.cell(row_num, 1, idx)
        for col_num, value in enumerate(row.tolist(), start=2):
            ws1.cell(row_num, col_num, int(value))
    end_row = start_row + len(pivot_df)
    style_table(ws1, start_row, 1, end_row, len(headers), grand_total_row=end_row)
    autosize_columns(ws1)
    ws1.freeze_panes = "A9"

    # Sheet 2
    ws2["A1"] = "Open vs Closed by Zone"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Business rule"
    ws2["A3"].font = Font(bold=True)
    ws2["A4"] = "Closed = Under review + Completed"
    ws2["A5"] = "Open = In progress + Not started"

    zone_headers = ["Zone", "Closed", "Open", "Grand Total"]
    zone_start_row = 8
    for col_num, header in enumerate(zone_headers, start=1):
        ws2.cell(zone_start_row, col_num, header)
    for row_num, zone in enumerate(ZONE_ORDER, start=zone_start_row + 1):
        ws2.cell(row_num, 1, zone)
        ws2.cell(row_num, 2, int(zone_summary.loc[zone, "Closed"]))
        ws2.cell(row_num, 3, int(zone_summary.loc[zone, "Open"]))
        ws2.cell(row_num, 4, int(zone_summary.loc[zone, "Grand Total"]))
    zone_end_row = zone_start_row + len(ZONE_ORDER)
    style_table(ws2, zone_start_row, 1, zone_end_row, len(zone_headers))
    autosize_columns(ws2)
    total_row = zone_end_row + 2
    ws2.cell(total_row, 1, "Total")
    ws2.cell(total_row, 2, int(zone_summary["Closed"].sum()))
    ws2.cell(total_row, 3, total_open)
    ws2.cell(total_row, 4, total_all)
    style_table(ws2, total_row, 1, total_row, 4)
    add_stacked_chart(ws2, "F3", f"{total_open}/{total_all} Open Assessment", zone_start_row, zone_end_row)
    ws2.freeze_panes = "A9"

    # Sheet 3
    ws3["A1"] = f"Overdue Assessment Summary (threshold = {overdue_threshold} days)"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3["A3"] = "Rule"
    ws3["A3"].font = Font(bold=True)
    ws3["A4"] = f"Overdue = Ageing >= {overdue_threshold} days"
    ws3["A5"] = "Closed overdue = Completed + Under review"
    ws3["A6"] = "Open overdue = In progress + Not started"

    od_headers = ["Zone", "Closed Overdue", "Open Overdue", "Overdue Total"]
    od_start_row = 9
    for col_num, header in enumerate(od_headers, start=1):
        ws3.cell(od_start_row, col_num, header)
    for row_num, zone in enumerate(ZONE_ORDER, start=od_start_row + 1):
        ws3.cell(row_num, 1, zone)
        ws3.cell(row_num, 2, int(overdue_zone.loc[zone, "Closed"]))
        ws3.cell(row_num, 3, int(overdue_zone.loc[zone, "Open"]))
        ws3.cell(row_num, 4, int(overdue_zone.loc[zone, "Grand Total"]))
    od_end_row = od_start_row + len(ZONE_ORDER)
    style_table(ws3, od_start_row, 1, od_end_row, len(od_headers))
    autosize_columns(ws3)
    od_total_row = od_end_row + 2
    ws3.cell(od_total_row, 1, "Total")
    ws3.cell(od_total_row, 2, int(overdue_zone["Closed"].sum()))
    ws3.cell(od_total_row, 3, total_overdue_open)
    ws3.cell(od_total_row, 4, total_overdue)
    style_table(ws3, od_total_row, 1, od_total_row, 4)
    add_stacked_chart(ws3, "F3", f"{total_overdue_open}/{total_overdue} Overdue Open Assessment ({overdue_threshold}+ days)", od_start_row, od_end_row)
    ws3.freeze_panes = "A10"

    wb.save(output_file)
    print(f"Created standalone output workbook: {output_file}")


if __name__ == "__main__":
    # Usage examples:
    # python onetrust_pivot_automation_v3.py input.xlsx
    # python onetrust_pivot_automation_v3.py input.xlsx output.xlsx
    # python onetrust_pivot_automation_v3.py input.xlsx output.xlsx 90
    in_file = sys.argv[1] if len(sys.argv) >= 2 else "PRP Sample Jun (2).xlsx"
    out_file = sys.argv[2] if len(sys.argv) >= 3 else None
    threshold = int(sys.argv[3]) if len(sys.argv) >= 4 else DEFAULT_OVERDUE_THRESHOLD
    build_reports(in_file, out_file, threshold)
