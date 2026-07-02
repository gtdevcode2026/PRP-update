import sys
from pathlib import Path

import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage


REQUIRED_SHEET = "OneTrust Assessment"
REQUIRED_COLUMNS = ["ID", "Stage", "Organization", "Ageing"]

STAGE_ORDER = ["Completed", "In progress", "Not started", "Under review"]

ORG_ORDER = [
    "Africa",
    "APAC",
    "BEES",
    "BEES | FINTECH",
    "Europe",
    "GHQ",
    "Middle America Zone",
    "North America Zone",
    "South America Zone",
]

ZONE_MAP = {
    "Africa": "AFR",
    "APAC": "APAC",
    "BEES": "GRO",
    "BEES | FINTECH": "GRO",
    "Europe": "EUR",
    "GHQ": "GHQ",
    "Middle America Zone": "MAZ",
    "North America Zone": "NAZ",
    "South America Zone": "SAZ",
}

ZONE_ORDER = ["AFR", "APAC", "GRO", "EUR", "GHQ", "MAZ", "NAZ", "SAZ"]

DEFAULT_OVERDUE_THRESHOLD = 90

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="EAF3FB")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def autosize_columns(ws, min_width=12, max_width=32):
    for col_cells in ws.columns:
        lengths = []

        for cell in col_cells:
            try:
                lengths.append(len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                lengths.append(0)

        width = max(lengths) + 2 if lengths else min_width
        width = max(min_width, min(max_width, width))

        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def style_table(ws, start_row, start_col, end_row, end_col, grand_total_row=None):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            cell.border = THIN_BORDER

            if row == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")

            elif grand_total_row and row == grand_total_row:
                cell.fill = SUBHEADER_FILL
                cell.font = HEADER_FONT

            if col > start_col:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(start_row + 1, end_row + 1):
        ws.cell(row, start_col).font = Font(bold=(grand_total_row == row))


def create_stacked_chart_png(
    chart_data,
    title,
    output_png,
    category_col="Zone",
    first_series="Closed",
    second_series="Open",
    first_label="Closed",
    second_label="Open",
    y_label="Assessment Count",
):
    zones = chart_data[category_col].tolist()
    first_values = chart_data[first_series].astype(int).tolist()
    second_values = chart_data[second_series].astype(int).tolist()

    x_positions = range(len(zones))

    plt.figure(figsize=(11, 5.5), dpi=140)

    bars_closed = plt.bar(
        x_positions,
        first_values,
        label=first_label,
        color="#2F75B5",
    )

    bars_open = plt.bar(
        x_positions,
        second_values,
        bottom=first_values,
        label=second_label,
        color="#ED7D31",
    )

    plt.title(title, fontsize=14, fontweight="bold", pad=14)
    plt.xlabel("Zone")
    plt.ylabel(y_label)

    plt.xticks(x_positions, zones, rotation=0)
    plt.legend(loc="upper center", bbox_to_anchor=(0.5, -0.12), ncol=2)

    plt.grid(axis="y", linestyle="--", alpha=0.3)

    max_total = max([a + b for a, b in zip(first_values, second_values)] + [1])
    plt.ylim(0, max_total * 1.20)

    for i, value in enumerate(first_values):
        if value > 0:
            plt.text(
                i,
                value / 2,
                str(value),
                ha="center",
                va="center",
                color="white",
                fontsize=9,
                fontweight="bold",
            )

    for i, value in enumerate(second_values):
        if value > 0:
            plt.text(
                i,
                first_values[i] + value / 2,
                str(value),
                ha="center",
                va="center",
                color="white",
                fontsize=9,
                fontweight="bold",
            )

    for i, total in enumerate([a + b for a, b in zip(first_values, second_values)]):
        if total > 0:
            plt.text(
                i,
                total + max_total * 0.03,
                str(total),
                ha="center",
                va="bottom",
                fontsize=9,
                fontweight="bold",
            )

    plt.tight_layout()
    plt.savefig(output_png, bbox_inches="tight")
    plt.close()


def insert_png_chart(ws, image_path, anchor_cell="F2", width=760, height=390):
    img = ExcelImage(str(image_path))
    img.width = width
    img.height = height
    ws.add_image(img, anchor_cell)


def prepare_chart_area(ws):
    for col in range(6, 18):
        ws.column_dimensions[get_column_letter(col)].width = 14

    for row in range(2, 24):
        ws.row_dimensions[row].height = 22


def build_reports(input_file, output_file=None, overdue_threshold=DEFAULT_OVERDUE_THRESHOLD):
    input_path = Path(input_file)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    if output_file is None:
        output_file = input_path.with_name("OneTrust_Report.xlsx")
    else:
        output_file = Path(output_file)

    df = pd.read_excel(
        input_path,
        sheet_name=REQUIRED_SHEET,
        engine="openpyxl",
    )

    df.columns = [str(col).strip() for col in df.columns]

    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]

    if missing:
        raise ValueError(f"Missing required columns in '{REQUIRED_SHEET}': {missing}")

    data = df[REQUIRED_COLUMNS].copy()

    data["Stage"] = data["Stage"].astype(str).str.strip()
    data["Organization"] = data["Organization"].astype(str).str.strip()
    data["Ageing"] = pd.to_numeric(data["Ageing"], errors="coerce")

    data = data[data["Stage"].isin(STAGE_ORDER)]
    data = data[data["Organization"].isin(ORG_ORDER)]

    data["Zone"] = data["Organization"].map(ZONE_MAP)
    data["Is Overdue"] = data["Ageing"] >= overdue_threshold

    pivot_df = pd.pivot_table(
        data,
        index="Organization",
        columns="Stage",
        values="ID",
        aggfunc="count",
        fill_value=0,
    )

    pivot_df = pivot_df.reindex(
        index=ORG_ORDER,
        columns=STAGE_ORDER,
        fill_value=0,
    )

    pivot_df["Grand Total"] = pivot_df.sum(axis=1)
    pivot_df.loc["Grand Total"] = pivot_df.sum(axis=0)

    zone_stage = pivot_df.drop(index="Grand Total").copy().reset_index()
    zone_stage["Zone"] = zone_stage["Organization"].map(ZONE_MAP)

    zone_summary = zone_stage.groupby("Zone", as_index=True).agg(
        {
            "Completed": "sum",
            "In progress": "sum",
            "Not started": "sum",
            "Under review": "sum",
            "Grand Total": "sum",
        }
    )

    zone_summary["Closed"] = (
        zone_summary["Completed"] + zone_summary["Under review"]
    )

    zone_summary["Open"] = (
        zone_summary["In progress"] + zone_summary["Not started"]
    )

    zone_summary = zone_summary.reindex(ZONE_ORDER).fillna(0)

    zone_summary = zone_summary[
        [
            "Closed",
            "Open",
            "Grand Total",
        ]
    ]

    total_closed = int(zone_summary["Closed"].sum())
    total_open = int(zone_summary["Open"].sum())
    total_all = int(zone_summary["Grand Total"].sum())

    overdue = data[data["Is Overdue"]].copy()

    overdue_zone = overdue.groupby(["Zone", "Stage"]).size().unstack(fill_value=0)

    overdue_zone = overdue_zone.reindex(
        index=ZONE_ORDER,
        columns=STAGE_ORDER,
        fill_value=0,
    ).fillna(0)

    overdue_zone["Closed"] = (
        overdue_zone["Completed"] + overdue_zone["Under review"]
    )

    overdue_zone["Open"] = (
        overdue_zone["In progress"] + overdue_zone["Not started"]
    )

    overdue_zone["Grand Total"] = overdue_zone["Closed"] + overdue_zone["Open"]

    total_overdue_closed = int(overdue_zone["Closed"].sum())
    total_overdue_open = int(overdue_zone["Open"].sum())
    total_overdue = int(overdue_zone["Grand Total"].sum())

    wb = Workbook()

    default_ws = wb.active
    wb.remove(default_ws)

    ws1 = wb.create_sheet("Auto Pivot Summary")
    ws2 = wb.create_sheet("Auto Open Closed")
    ws3 = wb.create_sheet(f"Auto Overdue {overdue_threshold}D")

    # ------------------------
    # Sheet 1: Pivot Summary
    # ------------------------
    ws1["A1"] = "Pivot-style summary from 'OneTrust Assessment'"
    ws1["A1"].font = Font(bold=True, size=14)

    ws1["A3"] = "Fields used"
    ws1["A3"].font = Font(bold=True)

    ws1["A4"] = "Rows"
    ws1["B4"] = "Organization"

    ws1["A5"] = "Columns"
    ws1["B5"] = "Stage"

    ws1["A6"] = "Values"
    ws1["B6"] = "Count of ID"

    start_row = 8
    headers = ["Organization"] + list(pivot_df.columns)

    for col_num, header in enumerate(headers, start=1):
        ws1.cell(start_row, col_num, header)

    for row_num, (idx, row) in enumerate(pivot_df.iterrows(), start=start_row + 1):
        ws1.cell(row_num, 1, idx)

        for col_num, value in enumerate(row.tolist(), start=2):
            ws1.cell(row_num, col_num, int(value))

    end_row = start_row + len(pivot_df)

    style_table(
        ws1,
        start_row,
        1,
        end_row,
        len(headers),
        grand_total_row=end_row,
    )

    autosize_columns(ws1)
    ws1.freeze_panes = "A9"

    # ------------------------
    # Sheet 2: Open vs Closed
    # ------------------------
    ws2["A1"] = "Open vs Closed by Zone"
    ws2["A1"].font = Font(bold=True, size=14)

    ws2["A3"] = "Business rule"
    ws2["A3"].font = Font(bold=True)

    ws2["A4"] = "Closed = Under review + Completed"
    ws2["A5"] = "Open = In progress + Not started"

    zone_headers = ["Zone", "Closed", "Open", "Grand Total"]
    zone_start_row = 8

    for col_num, header in enumerate(zone_headers, start=1):
        ws2.cell(zone_start_row, col_num, header)

    for row_num, zone in enumerate(ZONE_ORDER, start=zone_start_row + 1):
        ws2.cell(row_num, 1, zone)
        ws2.cell(row_num, 2, int(zone_summary.loc[zone, "Closed"]))
        ws2.cell(row_num, 3, int(zone_summary.loc[zone, "Open"]))
        ws2.cell(row_num, 4, int(zone_summary.loc[zone, "Grand Total"]))

    zone_end_row = zone_start_row + len(ZONE_ORDER)

    style_table(
        ws2,
        zone_start_row,
        1,
        zone_end_row,
        len(zone_headers),
    )

    total_row = zone_end_row + 2

    ws2.cell(total_row, 1, "Total")
    ws2.cell(total_row, 2, total_closed)
    ws2.cell(total_row, 3, total_open)
    ws2.cell(total_row, 4, total_all)

    style_table(ws2, total_row, 1, total_row, 4)

    autosize_columns(ws2)
    ws2.freeze_panes = "A9"

    # ------------------------
    # Sheet 3: Overdue Summary
    # ------------------------
    ws3["A1"] = f"Overdue Assessment Summary - Threshold = {overdue_threshold} days"
    ws3["A1"].font = Font(bold=True, size=14)

    ws3["A3"] = "Rule"
    ws3["A3"].font = Font(bold=True)

    ws3["A4"] = f"Overdue = Ageing >= {overdue_threshold} days"
    ws3["A5"] = "Closed overdue = Completed + Under review"
    ws3["A6"] = "Open overdue = In progress + Not started"

    od_headers = ["Zone", "Closed Overdue", "Open Overdue", "Overdue Total"]
    od_start_row = 9

    for col_num, header in enumerate(od_headers, start=1):
        ws3.cell(od_start_row, col_num, header)

    for row_num, zone in enumerate(ZONE_ORDER, start=od_start_row + 1):
        ws3.cell(row_num, 1, zone)
        ws3.cell(row_num, 2, int(overdue_zone.loc[zone, "Closed"]))
        ws3.cell(row_num, 3, int(overdue_zone.loc[zone, "Open"]))
        ws3.cell(row_num, 4, int(overdue_zone.loc[zone, "Grand Total"]))

    od_end_row = od_start_row + len(ZONE_ORDER)

    style_table(
        ws3,
        od_start_row,
        1,
        od_end_row,
        len(od_headers),
    )

    od_total_row = od_end_row + 2

    ws3.cell(od_total_row, 1, "Total")
    ws3.cell(od_total_row, 2, total_overdue_closed)
    ws3.cell(od_total_row, 3, total_overdue_open)
    ws3.cell(od_total_row, 4, total_overdue)

    style_table(ws3, od_total_row, 1, od_total_row, 4)

    autosize_columns(ws3)
    ws3.freeze_panes = "A10"

    # ------------------------
    # Matplotlib PNG Charts
    # ------------------------
    prepare_chart_area(ws2)
    prepare_chart_area(ws3)

    chart_files = []

    zone_chart_df = zone_summary.reset_index()
    zone_chart_df = zone_chart_df.rename(columns={"index": "Zone"})

    open_closed_png = output_file.with_name("open_closed_chart_temp.png")

    create_stacked_chart_png(
        chart_data=zone_chart_df,
        title=f"{total_open}/{total_all} Open Assessment",
        output_png=open_closed_png,
        category_col="Zone",
        first_series="Closed",
        second_series="Open",
        first_label="Closed",
        second_label="Open",
        y_label="Assessment Count",
    )

    chart_files.append(open_closed_png)

    insert_png_chart(
        ws=ws2,
        image_path=open_closed_png,
        anchor_cell="F2",
        width=760,
        height=390,
    )

    overdue_chart_df = overdue_zone[["Closed", "Open", "Grand Total"]].copy()
    overdue_chart_df = overdue_chart_df.reset_index()
    overdue_chart_df = overdue_chart_df.rename(columns={"index": "Zone"})

    overdue_png = output_file.with_name("overdue_chart_temp.png")

    create_stacked_chart_png(
        chart_data=overdue_chart_df,
        title=f"{total_overdue_open}/{total_overdue} Overdue Open Assessment ({overdue_threshold}+ days)",
        output_png=overdue_png,
        category_col="Zone",
        first_series="Closed",
        second_series="Open",
        first_label="Closed Overdue",
        second_label="Open Overdue",
        y_label="Overdue Assessment Count",
    )

    chart_files.append(overdue_png)

    insert_png_chart(
        ws=ws3,
        image_path=overdue_png,
        anchor_cell="F2",
        width=760,
        height=390,
    )

    # Save workbook
    wb.save(output_file)

    # Delete temporary PNGs after embedding
    for chart_file in chart_files:
        try:
            Path(chart_file).unlink()
        except Exception:
            pass

    print(f"Created standalone output workbook with embedded PNG charts: {output_file}")


if __name__ == "__main__":
    # Usage examples:
    # python onetrust_pivot_automation_matplotlib.py "PRP Sample Jun (2).xlsx"
    # python onetrust_pivot_automation_matplotlib.py "PRP Sample Jun (2).xlsx" "OneTrust_Report.xlsx"
    # python onetrust_pivot_automation_matplotlib.py "PRP Sample Jun (2).xlsx" "OneTrust_Report.xlsx" 90

    in_file = sys.argv[1] if len(sys.argv) >= 2 else "PRP Sample Jun (2).xlsx"
    out_file = sys.argv[2] if len(sys.argv) >= 3 else None
    threshold = int(sys.argv[3]) if len(sys.argv) >= 4 else DEFAULT_OVERDUE_THRESHOLD

    build_reports(in_file, out_file, threshold)