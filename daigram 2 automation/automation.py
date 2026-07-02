import os
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image


# =========================
# INPUT / OUTPUT
# =========================
input_file = r"PRP Sample Jun (2).xlsx"
sheet_name = "OneTrust Assessment"
output_file = "output file D2.xlsx"

chart1_file = "chart_2026_assessment_status.png"
chart2_file = "chart_supplier_response_time.png"


# =========================
# CHECK FILE
# =========================
if not os.path.exists(input_file):
    raise FileNotFoundError(f"Input file not found: {input_file}")


# =========================
# LOAD DATA
# =========================
df = pd.read_excel(input_file, sheet_name=sheet_name, engine="openpyxl")
df.columns = df.columns.str.strip()

required_columns = ["ID", "Organization", "Stage", "Date created", "Tags"]
missing = [c for c in required_columns if c not in df.columns]

if missing:
    raise ValueError(f"Missing required columns: {missing}")

for c in ["Organization", "Stage", "Tags"]:
    df[c] = df[c].fillna("").astype(str).str.strip()

df["Date created"] = pd.to_datetime(df["Date created"], errors="coerce")


# =========================
# FILTERS
# =========================
filtered = df[
    df["Tags"].str.contains(r"(?i)^cyber$", na=False)
    & (df["Date created"].dt.year == 2026)
].copy()


# =========================
# STAGE LOGIC
# =========================
closed_stages = {"Completed", "Under review"}

filtered["Final Stage"] = filtered["Stage"].apply(
    lambda x: "Closed" if str(x).strip() in closed_stages else "Open"
)


# =========================
# ORG MAPPING
# =========================
org_map = {
    "Africa": "AFR",
    "APAC": "APAC",
    "BEES": "GRO",
    "BEES | FINTECH": "GRO",
    "Europe": "Europe",
    "GHQ": "GHQ",
    "South America Zone": "SAZ",
    "North America Zone": "NAZ",
    "Middle America Zone": "MAZ",
}

filtered["Org Display"] = filtered["Organization"].map(
    lambda x: org_map.get(x, x)
)


# =========================
# PIVOT TABLE
# =========================
pivot = pd.pivot_table(
    filtered,
    index="Org Display",
    columns="Final Stage",
    values="ID",
    aggfunc="count",
    fill_value=0,
)

for col in ["Open", "Closed"]:
    if col not in pivot.columns:
        pivot[col] = 0

pivot = pivot[["Open", "Closed"]].reset_index()
pivot.columns = ["Row Labels", "Open", "Closed"]
pivot["Grand Total"] = pivot["Open"] + pivot["Closed"]

order = ["AFR", "APAC", "GRO", "Europe", "GHQ", "SAZ", "MAZ", "NAZ"]

pivot["sort_order"] = pivot["Row Labels"].apply(
    lambda x: order.index(x) if x in order else 999
)

pivot = (
    pivot
    .sort_values(["sort_order", "Row Labels"])
    .drop(columns="sort_order")
    .reset_index(drop=True)
)

grand_total_row = pd.DataFrame([
    {
        "Row Labels": "Grand Total",
        "Open": int(pivot["Open"].sum()),
        "Closed": int(pivot["Closed"].sum()),
        "Grand Total": int(pivot["Grand Total"].sum()),
    }
])

pivot_display = pd.concat([pivot, grand_total_row], ignore_index=True)


# =========================
# KPI DATA
# =========================
baseline_25 = 0.60
q1_26 = 0.32
target_26 = 0.65

closed_total = int(filtered["Final Stage"].eq("Closed").sum())
record_total = int(len(filtered))

q2_26 = round(closed_total / record_total, 2) if record_total else 0

kpi_data = [
    ("Baseline '25", baseline_25, "static"),
    ("Q1 '26", q1_26, "static"),
    ("Q2 '26", q2_26, ""),
    ("Target '26", target_26, "static"),
]


# =====================================================
# CREATE CHART 1 USING MATPLOTLIB
# =====================================================
chart_df = pivot.copy()

orgs = chart_df["Row Labels"].tolist()
open_values = chart_df["Open"].astype(int).tolist()
closed_values = chart_df["Closed"].astype(int).tolist()

x = list(range(len(orgs)))

fig, ax = plt.subplots(figsize=(7.6, 4.4), dpi=180)

fig.patch.set_facecolor("black")
ax.set_facecolor("black")

open_color = "#00AEEF"
closed_color = "#D4AF37"

bars_open = ax.bar(
    x,
    open_values,
    color=open_color,
    width=0.48,
    label="Open"
)

bars_closed = ax.bar(
    x,
    closed_values,
    bottom=open_values,
    color=closed_color,
    width=0.48,
    label="Closed"
)

ax.set_title(
    f"2026 Assessment\n({closed_total}/{record_total})",
    color="white",
    fontsize=16,
    fontweight="bold",
    pad=18
)

ax.set_xticks(x)
ax.set_xticklabels(
    orgs,
    rotation=45,
    ha="right",
    color="white",
    fontsize=11
)

ax.tick_params(axis="y", left=False, labelleft=False)
ax.tick_params(axis="x", colors="white")

ax.grid(False)

for spine in ["top", "right", "left"]:
    ax.spines[spine].set_visible(False)

ax.spines["bottom"].set_color("white")
ax.axhline(0, color="white", linewidth=1)

# Labels inside bars
for i, value in enumerate(open_values):
    if value > 0:
        ax.text(
            x[i],
            value / 2,
            str(value),
            ha="center",
            va="center",
            color="white",
            fontsize=10,
            fontweight="bold"
        )

for i, value in enumerate(closed_values):
    if value > 0:
        ax.text(
            x[i],
            open_values[i] + value / 2,
            str(value),
            ha="center",
            va="center",
            color="white",
            fontsize=10,
            fontweight="bold"
        )

# Legend
handles, labels = ax.get_legend_handles_labels()
legend_order = [1, 0]  # Closed first, Open second

legend = ax.legend(
    [handles[i] for i in legend_order],
    [labels[i] for i in legend_order],
    loc="upper center",
    bbox_to_anchor=(0.5, 0.88),
    ncol=2,
    frameon=False,
    fontsize=11
)

for text in legend.get_texts():
    text.set_color("white")

max_total = max([o + c for o, c in zip(open_values, closed_values)] + [1])
ax.set_ylim(0, max_total * 1.35)

plt.tight_layout()
plt.savefig(chart1_file, facecolor=fig.get_facecolor(), bbox_inches="tight")
plt.close(fig)


# =====================================================
# CREATE CHART 2 USING MATPLOTLIB
# =====================================================
metrics = [item[0] for item in kpi_data]
values = [item[1] for item in kpi_data]

fig, ax = plt.subplots(figsize=(7.6, 3.8), dpi=180)

bars = ax.barh(
    metrics,
    values,
    color="#4472C4",
    height=0.48
)

ax.set_title(
    "Improve in Supplier Response Time",
    fontsize=15,
    fontweight="bold",
    pad=14
)

ax.set_xlim(0, 1)

ax.set_xticks([0, 0.25, 0.50, 0.75, 1.00])
ax.set_xticklabels(["0%", "25%", "50%", "75%", "100%"])

ax.invert_yaxis()

for bar, value in zip(bars, values):
    ax.text(
        value + 0.02,
        bar.get_y() + bar.get_height() / 2,
        f"{value:.0%}",
        va="center",
        ha="left",
        fontsize=11,
        fontweight="bold"
    )

ax.spines["top"].set_visible(False)
ax.spines["right"].set_visible(False)

ax.grid(axis="x", linestyle="--", alpha=0.35)

plt.tight_layout()
plt.savefig(chart2_file, bbox_inches="tight")
plt.close(fig)


# =========================
# CREATE EXCEL
# =========================
wb = Workbook()
ws = wb.active
ws.title = "Dashboard"


# =========================
# STYLES
# =========================
header_fill = PatternFill("solid", fgColor="B7DEE8")
subheader_fill = PatternFill("solid", fgColor="D9EAF7")
grand_fill = PatternFill("solid", fgColor="B7DEE8")

bold = Font(bold=True)
thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")


# =========================
# FILTER AREA
# =========================
ws["A1"] = "Tags"
ws["B1"] = "Cyber"
ws["A2"] = "Date created"
ws["B2"] = "2026"

for cell in ["A1", "A2"]:
    ws[cell].fill = header_fill
    ws[cell].font = bold
    ws[cell].border = border

for cell in ["B1", "B2"]:
    ws[cell].fill = subheader_fill
    ws[cell].border = border


# =========================
# PIVOT TABLE OUTPUT
# =========================
start_row = 4
headers = list(pivot_display.columns)

for col_idx, h in enumerate(headers, start=1):
    c = ws.cell(row=start_row, column=col_idx, value=h)
    c.fill = header_fill
    c.font = bold
    c.border = border
    c.alignment = center

for r_idx, row in enumerate(pivot_display.itertuples(index=False), start=start_row + 1):
    for c_idx, value in enumerate(row, start=1):
        c = ws.cell(row=r_idx, column=c_idx, value=value)
        c.border = border

        if c_idx > 1:
            c.alignment = center

    if row[0] == "Grand Total":
        for c_idx in range(1, len(headers) + 1):
            ws.cell(row=r_idx, column=c_idx).fill = grand_fill
            ws.cell(row=r_idx, column=c_idx).font = bold


# =========================
# KPI TABLE
# =========================
kpi_start_row = start_row + len(pivot_display) + 5

for idx, text in enumerate(["Metric", "Value", "Remark"], start=1):
    c = ws.cell(row=kpi_start_row, column=idx, value=text)
    c.fill = header_fill
    c.font = bold
    c.border = border
    c.alignment = center

for i, (metric, value, remark) in enumerate(kpi_data, start=kpi_start_row + 1):
    ws.cell(row=i, column=1, value=metric)
    ws.cell(row=i, column=2, value=value)
    ws.cell(row=i, column=3, value=remark)

    for col in [1, 2, 3]:
        ws.cell(row=i, column=col).border = border

    ws.cell(row=i, column=2).number_format = "0%"

note_row = kpi_start_row + len(kpi_data) + 2

ws.cell(row=note_row, column=1, value="Q2 formula")
ws.cell(
    row=note_row,
    column=2,
    value=f"Closed / Grand Total = {closed_total} / {record_total}"
)

ws.cell(row=note_row, column=1).font = bold


# =========================
# INSERT CHART IMAGES
# =========================
img1 = Image(chart1_file)
img1.width = 560
img1.height = 360
ws.add_image(img1, "F2")

img2 = Image(chart2_file)
img2.width = 560
img2.height = 300
ws.add_image(img2, "F23")


# =========================
# FORMATTING
# =========================
widths = {
    "A": 18,
    "B": 12,
    "C": 12,
    "D": 14,
    "E": 4,
    "F": 14,
    "G": 14,
    "H": 14,
    "I": 14,
    "J": 14,
    "K": 14,
    "L": 14,
    "M": 14,
}

for col, width in widths.items():
    ws.column_dimensions[col].width = width

for row in range(1, 50):
    ws.row_dimensions[row].height = 22

ws.freeze_panes = "A5"


# =========================
# SAVE FILE
# =========================
wb.save(output_file)

print("SUCCESS")
print(output_file)