import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# =========================
# STEP 1: AUTO PICK FILE
# =========================
file_path = r"C:\Users\C915662\OneDrive - Anheuser-Busch InBev\AB-InBev Automations\Assessment_Automation\PRP Sample Jun (2).xlsx"


# =========================

# STEP 2: READ SHEET
# =========================
df = pd.read_excel(file_path, sheet_name="OneTrust Assessment", engine="openpyxl")
df.columns = df.columns.str.strip()

# =========================
# STEP 3: COLUMN DETECTION
# =========================
stage_col = [c for c in df.columns if "Stage" in c][0]
org_col = [c for c in df.columns if "Organization" in c][0]
working1_col = [c for c in df.columns if "Working1" in c][0]
working2_col = [c for c in df.columns if "Working2" in c][0]

# =========================
# STEP 4: APPLY FILTERS
# =========================
df_filtered = df[
    (df[working2_col] == "Beyond 1 Year Overdue") &
    (df[working1_col].isin(["Pending", "Completed in 2026"]))
].copy()

# =========================
# STEP 5: STAGE MAPPING
# =========================
def classify_stage(x):
    x = str(x).strip()
    if x in ["Under review", "Completed"]:
        return "Completed"
    elif x in ["Not started", "In progress"]:
        return "Open"
    else:
        return "Open"

df_filtered["Final Status"] = df_filtered[stage_col].apply(classify_stage)

# =========================
# STEP 6: CREATE PIVOT
# =========================
pivot = pd.pivot_table(
    df_filtered,
    index=org_col,
    columns=working1_col,
    values="Final Status",
    aggfunc="count",
    fill_value=0
)

# Add totals
pivot["Grand Total"] = pivot.sum(axis=1)
pivot.loc["Grand Total"] = pivot.sum()

# =========================
# STEP 7: SAVE TO EXCEL
# =========================
output_path = r"C:\Users\C915662\OneDrive - Anheuser-Busch InBev\AB-InBev Automations\Assessment_Automation\output.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    pivot.to_excel(writer, sheet_name="Summary")

# =========================
# =========================
# STEP 8: ADD STACKED BAR CHART (LIKE YOUR IMAGE)
# =========================
wb = load_workbook(output_path)
ws = wb["Summary"]

chart = BarChart()
chart.type = "col"
chart.grouping = "stacked"   # ✅ KEY: stacked chart
chart.overlap = 100          # ✅ full stacking

chart.title = "Assessments Completed vs Open"
chart.y_axis.title = "Count"
chart.x_axis.title = "Region"

# Data (ONLY Pending + Completed in 2026, exclude Grand Total column)
data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row-1)
cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row-1)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

# =========================
# ADD COLORS (MATCH YOUR IMAGE)
# =========================
series = chart.series

# Completed in 2026 → RED
series[0].graphicalProperties.solidFill = "FF0000"

# Pending → ORANGE/YELLOW
series[1].graphicalProperties.solidFill = "FFC000"

# =========================
# ADD CHART TO SHEET
# =========================
ws.add_chart(chart, "H5")

wb.save(output_path)