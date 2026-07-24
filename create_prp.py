import os
from datetime import datetime
from openpyxl import load_workbook, Workbook

# =====================================================
# PATHS
# =====================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

INPUT_FOLDER = os.path.join(BASE_DIR, "input")
OUTPUT_FOLDER = os.path.join(BASE_DIR, "output")

OUTPUT_FILE = "PRP Sample Jun (2).xlsx"

# =====================================================
# FIND FILES
# =====================================================

def find_source_files():

    risk_file = None
    assessment_file = None
    tprm_file = None

    print("\nScanning Input Folder...\n")

    if not os.path.exists(INPUT_FOLDER):
        raise Exception(
            f"Input folder not found:\n{INPUT_FOLDER}"
        )

    for file in os.listdir(INPUT_FOLDER):

        lower = file.lower()

        print(f"Found: {file}")

        if not lower.endswith(".xlsx"):
            continue

        if "risk-export" in lower:
            risk_file = os.path.join(INPUT_FOLDER, file)

        elif "assessment-export" in lower:
            assessment_file = os.path.join(INPUT_FOLDER, file)

        elif "tprm" in lower or "supplier" in lower:
            tprm_file = os.path.join(INPUT_FOLDER, file)

    print("\nDetected Files")
    print("-" * 50)
    print("Risk File       :", risk_file)
    print("Assessment File :", assessment_file)
    print("TPRM File       :", tprm_file)

    return risk_file, assessment_file, tprm_file


# =====================================================
# COPY SHEET
# =====================================================

def copy_sheet(source_ws, target_ws):

    for row in source_ws.iter_rows():

        for cell in row:

            target_ws[cell.coordinate] = cell.value


# =====================================================
# FIND COLUMN
# =====================================================

def find_column(ws, header_name):

    for col in range(1, ws.max_column + 1):

        value = ws.cell(1, col).value

        if value is not None:

            if str(value).strip().lower() == header_name.lower():

                return col

    return None


# =====================================================
# DATE CONVERTER
# =====================================================

def convert_date_column(ws, col_num):

    if not col_num:
        return

    formats = [
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%m/%d/%Y",
        "%Y-%m-%d",
        "%m-%d-%Y"
    ]

    for row in range(2, ws.max_row + 1):

        cell = ws.cell(row=row, column=col_num)

        if cell.value is None:
            continue

        if isinstance(cell.value, datetime):
            cell.number_format = "dd-mm-yyyy"
            continue

        value = str(cell.value).strip()

        for fmt in formats:

            try:

                cell.value = datetime.strptime(
                    value,
                    fmt
                )

                cell.number_format = "dd-mm-yyyy"

                break

            except:
                pass


# =====================================================
# ADD COLUMN
# =====================================================

def add_column(ws, header_name):

    for col in range(1, ws.max_column + 1):

        value = ws.cell(1, col).value

        if value is not None:

            if str(value).strip() == header_name:

                return col

    new_col = ws.max_column + 1

    ws.cell(
        row=1,
        column=new_col
    ).value = header_name

    return new_col


# =====================================================
# RISK EXPORT
# =====================================================

def process_risk_export(ws):

    print("Processing Risk Export")

    date_created = find_column(
        ws,
        "Date created"
    )

    convert_date_column(
        ws,
        date_created
    )

    aging_col = add_column(
        ws,
        "Aging"
    )

    for row in range(2, ws.max_row + 1):

        ws.cell(
            row=row,
            column=aging_col
        ).value = (
            f"=NETWORKDAYS(N{row},TODAY())"
        )


# =====================================================
# ASSESSMENT
# =====================================================

def process_assessment(ws):

    print("Processing Assessment")

    date_created = find_column(
        ws,
        "Date created"
    )

    date_submitted = find_column(
        ws,
        "Date submitted"
    )

    convert_date_column(
        ws,
        date_created
    )

    convert_date_column(
        ws,
        date_submitted
    )

    ageing_col = add_column(
        ws,
        "Ageing"
    )

    working1_col = add_column(
        ws,
        "Working1"
    )

    working2_col = add_column(
        ws,
        "Working2"
    )

    sep26_col = add_column(
        ws,
        "30sep26"
    )

    for row in range(2, ws.max_row + 1):

        # Ageing

        ws.cell(
            row=row,
            column=ageing_col
        ).value = (
            f"=NETWORKDAYS(N{row},TODAY())"
        )

        # Working1

        ws.cell(
            row=row,
            column=working1_col
        ).value = (
            f'=IF(AND(OR(C{row}="Completed",'
            f'C{row}="Under Review"),'
            f'O{row}>DATE(2025,12,31)),'
            f'"Completed in 2026",'
            f'IF(O{row}="",'
            f'"Pending",'
            f'"Completed Before 2026"))'
        )

        # Working2

        ws.cell(
            row=row,
            column=working2_col
        ).value = (
            f'=IF(DATE(2026,9,30)-N{row}>=365,'
            f'"Beyond 1 Year Overdue",'
            f'"Current")'
        )

        # 30sep26

        ws.cell(
            row=row,
            column=sep26_col
        ).value = (
            f"=DATE(2026,9,30)-N{row}"
        )


# =====================================================
# MAIN
# =====================================================

def main():

    risk_file, assessment_file, tprm_file = (
        find_source_files()
    )

    if not risk_file:
        raise Exception(
            "Risk export file not found."
        )

    if not assessment_file:
        raise Exception(
            "Assessment export file not found."
        )

    if not tprm_file:
        raise Exception(
            "TPRM export file not found."
        )

    wb_new = Workbook()

    default_sheet = wb_new.active
    wb_new.remove(default_sheet)

    ws_risk = wb_new.create_sheet(
        "OneTrust - Risk Export"
    )

    ws_tprm = wb_new.create_sheet(
        "TPRM WebApp Portal"
    )

    ws_assessment = wb_new.create_sheet(
        "OneTrust Assessment"
    )

    print("\nLoading Risk File")
    wb_risk = load_workbook(risk_file)
    copy_sheet(
        wb_risk.active,
        ws_risk
    )

    print("Loading TPRM File")
    wb_tprm = load_workbook(tprm_file)
    copy_sheet(
        wb_tprm.active,
        ws_tprm
    )

    print("Loading Assessment File")
    wb_assessment = load_workbook(
        assessment_file
    )
    copy_sheet(
        wb_assessment.active,
        ws_assessment
    )

    process_risk_export(
        ws_risk
    )

    process_assessment(
        ws_assessment
    )

    os.makedirs(
        OUTPUT_FOLDER,
        exist_ok=True
    )

    output_path = os.path.join(
        OUTPUT_FOLDER,
        OUTPUT_FILE
    )

    wb_new.save(
        output_path
    )

    print("\nSUCCESS")
    print(f"\nOutput Created:\n{output_path}")


if __name__ == "__main__":
    main()